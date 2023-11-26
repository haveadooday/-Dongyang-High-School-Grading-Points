
## 성적 수정 제대로 안되는 부분 수정
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from openpyxl import Workbook, load_workbook
import logging
from datetime import datetime
import hashlib

logging.basicConfig(filename = 'logs/'+datetime.today().strftime("%Y.%m.%d")+'.log', level = logging.DEBUG)
app = Flask(__name__)
app.secret_key = 'Raeto0322'

excel_file = 'students.xlsx'
excel_file_2 = 'students2.xlsx'

def hasher(password):
    hash_object = hashlib.sha512()
    hash_object.update(password.encode('utf-8'))
    return hash_object.hexdigest()

def validate_score(score):
    if score is None:
        return True
    else:
        return 0 <= score <= 100

def calculate_cut_score(people, scores):
    scores = [s for s in scores if s is not None]
    count = len(scores) 
    sorted_scores = sorted(scores, reverse=True)
    percent = [4,11,23,40,60,77,89,96,100]
    percentile_index = []
    if count >= 25:
        for i in percent:
            if count > 0:  
                percentile_position = round(count * (i / 100))
                index = min(percentile_position - 1, count - 1)
                percentile_index.append(sorted_scores[index])
            else:
                percentile_index.append(None)

        return percentile_index
    else:
        percentile_index = []
        for i in range(10):
            percentile_index.append(None)
        return percentile_index

def cap_score_range(score):
    return max(0, min(score, 100))

def calculate_score_range(scores):
    ranges = []
    for score_list in scores:
        score_range = calculate_cut_score(len(score_list), score_list)
        score_range = [f"{max(0, score) if score is not None else '산출중'}-{cap_score_range(score+1) if score is not None else '산출중'}" for score in score_range]
        ranges.append(score_range)
    return ranges

@app.route('/download/dongyang.ico')
def download_dongyang_icon():
    directory = 'static/'  # 실제 파일이 위치하는 디렉토리 경로를 입력해주세요
    return send_file(f'{directory}/dongyang.ico', as_attachment=True)

@app.after_request
def add_security_headers(response):
    # X-Frame-Options 헤더 추가 (클릭재킹 방지)
    response.headers['X-Frame-Options'] = 'DENY'
    
    # X-Content-Type-Options 헤더 추가 (MIME 타입 스니핑 방지)
    response.headers['X-Content-Type-Options'] = 'nosniff'

    return response

@app.route('/privacy')
def privacy():
    return render_template('privacy.html')

@app.route('/modify')
def modify():
    return redirect(url_for('grade_modify_route'))
    #return render_template('modify_student.html',student_data={})

@app.route('/add_student_form_1', methods=['POST'])
def add_or_update_student_route():
    return add_student()

@app.route('/add_student_form_2_lib', methods=['POST'])
def add_student_route_2_lib():
    return add_student_2_lib()

@app.route('/add_student_form_2_science', methods=['POST'])
def add_student_route_2_science():
    return add_student_2_science()


@app.route('/modify_student_1', methods=['POST'])
def update_student_route():
    action = request.form['action']  
    if action == '수정':
        return modify_student()
    
    elif action == '불러오기':
        student_data = load_student()
        if student_data:
            return render_template('modify_student.html', student_data=student_data)
        else:
            flash('학생 정보를 찾을 수 없습니다.')
            return render_template('modify_student.html', student_data={})

    else:
        return "알 수 없는 작업입니다."

@app.route('/modify_student_2_lib', methods=['POST'])
def update_student_route_2_lib():
    action = request.form['action']  
    if action == '수정':
        return modify_student_2_lib()
    
    elif action == '불러오기':
        student_data = load_student_2()
        if student_data:
            return render_template('modify_student_2_lib.html', student_data=student_data)
        else:
            flash('학생 정보를 찾을 수 없습니다.')
            return render_template('modify_student_2_lib.html', student_data={})

    else:
        return "알 수 없는 작업입니다."

@app.route('/modify_student_2_science', methods=['POST'])
def update_student_route_2_science():
    action = request.form['action']  
    if action == '수정':
        return modify_student_2_science()
    
    elif action == '불러오기':
        student_data = load_student_2()
        if student_data:
            return render_template('modify_student_2_science.html', student_data=student_data)
        else:
            flash('학생 정보를 찾을 수 없습니다.')
            return render_template('modify_student_2_science.html', student_data={})

    else:
        return "알 수 없는 작업입니다."

@app.route('/static/<path:path>')
def send_static(path):
    return send_from_directory('static', path, mimetype='font/ttf')

@app.route('/')
def index():
    #return render_template('index.html')
    return render_template('seeyousoon.html')

@app.route('/add_student_form')
def add_student_form():
    return redirect(url_for('grade_add_student_route'))

@app.route('/add_student_form_1')
def add_student_form_1():
    return render_template('add_student.html')

@app.route('/add_student_form_2_lib')
def add_student_form_2_lib():
    return render_template('add_student_2_lib.html')

@app.route('/add_student_form_2_science')
def add_student_form_2_science():
    return render_template('add_student_2_science.html')

@app.route('/grade/add_student')
def grade_add_student_route():
    return render_template('grade.html')

@app.route('/grade/modify')
def grade_modify_route():
    return render_template('grade_modify.html')

@app.route('/grade/average')
def grade_average_route():
    return render_template('grade_average.html')


@app.route('/modify_student_form_1')
def modify_student_form_1():
    return render_template('modify_student.html', student_data={})

@app.route('/modify_student_form_2_lib')
def modify_student_form_2_lib():
    return render_template('modify_student_2_lib.html', student_data={})

@app.route('/modify_student_form_2_science')
def modify_student_form_2_science():
    return render_template('modify_student_2_science.html', student_data={})

@app.route('/depart/modify')
def depart_modify_student():
    return render_template('lib_science_modify.html')

@app.route('/depart/add_student')
def depart_add_student():
    return render_template('lib_science.html')

@app.route('/grade/<type1>', methods=['POST'])
def grade_route(type1):
    action = request.form['action']
    if action == '1학년':
        if type1 == 'add_student':
            return redirect(url_for('add_student_form_1'))
        elif type1 == 'modify':
            return redirect(url_for('modify_student_form_1'))
        elif type1 == 'average':
            return redirect(url_for('show_cut_scores'))
    elif action == '2학년':
        if type1 == 'add_student':
            return redirect(url_for('depart_add_student'))
        elif type1 == 'modify':
            return redirect(url_for('depart_modify_student'))
        elif type1 == 'average':
            return redirect(url_for('show_cut_scores_2'))
    else:
        return "알 수 없는 작업입니다."
    
@app.route('/depart/<type1>', methods=['POST'])
def depart_route(type1):
    action = request.form['action']
    if action == '문과':
        if type1 == 'add_student':
            return redirect(url_for('add_student_form_2_lib'))
        elif type1 == 'modify':
            return redirect(url_for('modify_student_form_2_lib'))
    elif action == '이과':
        if type1 == 'add_student':
            return redirect(url_for('add_student_form_2_science'))
        elif type1 == 'modify':
            return redirect(url_for('modify_student_form_2_science'))
    else:
        return "알 수 없는 작업입니다."
    
def modify_student_2_lib():
    name = hasher(request.form.get('name'))
    password = hasher(request.form.get('password'))
    subjects = ['read', 'math2', 'english2', 'korean_geo', 'east_history', 'eco', 'law', 'japanese']
    subject_values = [float(request.form.get(subject)) if request.form.get(subject).strip() else None for subject in subjects]
    physical = chemisty = life = None
    for subject in subjects:
        value = request.form.get(subject)
        if value.strip():
            globals()[subject] = float(value)
        else:
            globals()[subject] = None
            
    if not all(validate_score(score) for score in subject_values):
        flash('성적은 0부터 100 사이어야 합니다.')
    else:
        try:
            wb = load_workbook(excel_file_2)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I'])

        found_index = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == name and row[1] == password:
                found_index = idx  # 찾은 튜플의 인덱스 저장
                break
        if found_index is not None:
            column_headers = ['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I']
            subjects = [read, math2, english2,physical, chemisty, life, korean_geo, east_history, eco, law, japanese]
            i = 3
            for subjects in subjects:
                if subjects is not None:
                    ws.cell(row=found_index,column=i,value=subjects)
                else:
                    ws.cell(row=found_index,column=i,value='')
                i += 1

            flash('성적이 수정되었습니다.')

            wb.save(excel_file_2)
        else:
            flash('성적을 찾지 못했습니다.')

    return redirect(url_for('modify_student_form_2_lib'))


def modify_student_2_science():
    name = hasher(request.form.get('name'))
    password = hasher(request.form.get('password'))
    subjects = ['read', 'math2', 'english2','physical', 'chemisty','life' , 'japanese']
    subject_values = [float(request.form.get(subject)) if request.form.get(subject).strip() else None for subject in subjects]
    korean_geo = east_history = eco = law = None
    for subject in subjects:
        value = request.form.get(subject)
        if value.strip():
            globals()[subject] = float(value)
        else:
            globals()[subject] = None
            
    if not all(validate_score(score) for score in subject_values):
        flash('성적은 0부터 100 사이어야 합니다.')
    else:
        try:
            wb = load_workbook(excel_file_2)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I'])

        found_index = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == name and row[1] == password:
                found_index = idx  # 찾은 튜플의 인덱스 저장
                break
        if found_index is not None:
            column_headers = ['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I']
            subjects = [read, math2, english2,physical, chemisty, life, korean_geo, east_history, eco, law, japanese]
            i = 3
            for subjects in subjects:
                if subjects is not None:
                    ws.cell(row=found_index,column=i,value=subjects)
                else:
                    ws.cell(row=found_index,column=i,value='')
                i += 1

            flash('성적이 수정되었습니다.')

            wb.save(excel_file_2)
        else:
            flash('성적을 찾지 못했습니다.')

    return redirect(url_for('modify_student_form_2_science'))

def modify_student():
    name = hasher(request.form.get('name'))
    password = hasher(request.form.get('password'))
    subjects = ['korean', 'math', 'english', 'social', 'science', 'korean_history']
    subject_values = [float(request.form.get(subject)) if request.form.get(subject).strip() else None for subject in subjects]
    for subject in subjects:
        value = request.form.get(subject)
        if value.strip():
            globals()[subject] = float(value)
        else:
            globals()[subject] = None
            
    if not all(validate_score(score) for score in subject_values):
        flash('성적은 0부터 100 사이어야 합니다.')
    else:
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['이름', '비밀번호', '국어', '수학', '영어','사회','과학','한국사'])

        found_index = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == name and row[1] == password:
                found_index = idx  # 찾은 튜플의 인덱스 저장
                break
        if found_index is not None:
            column_headers = ["이름", "비밀번호", "국어", "수학", "영어", "사회", "과학", "한국사"]
            subjects = [korean, math, english, social, science, korean_history]
            i = 3
            for subjects in subjects:
                if subjects is not None:
                    ws.cell(row=found_index,column=i,value=subjects)
                else:
                    ws.cell(row=found_index,column=i,value='')
                i += 1

            flash('성적이 수정되었습니다.')

            wb.save(excel_file)
        else:
            flash('성적을 찾지 못했습니다.')

    return redirect(url_for('modify_student_form_1'))

def load_student():
    name = request.form.get('name')
    password = request.form.get('password')
    wb = load_workbook(excel_file)
    ws = wb.active
    
    student_info = None
                                                    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(hasher(name)) and str(row[1]) == str(hasher(password)):
            student_info = {
                '이름': name,
                '비밀번호': password,
                '국어': row[2],
                '수학': row[3],
                '영어': row[4],
                '사회': row[5],
                '과학': row[6],
                '한국사': row[7]
            }
            break
        
    if student_info is None:
        return None

    return student_info

def load_student_2():
    name = request.form.get('name')
    password = request.form.get('password')
    wb = load_workbook(excel_file_2)
    ws = wb.active
    
    student_info = None
                                                    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(hasher(name)) and str(row[1]) == str(hasher(password)):
            student_info = {
                '이름': name,
                '비밀번호': password,
                '국어': row[2],
                '수학': row[3],
                '영어': row[4],
                '물리': row[5],
                '화학': row[6],
                '생명': row[7],
                '한국지리': row[8],
                '동아시아사': row[9],
                '경제': row[10],
                '정치와법': row[11],
                '일본어': row[12],
            }
            break
        
    if student_info is None:
        return None

    return student_info

def add_student_2_science():
    name = hasher(request.form.get('name'))
    password = hasher(request.form.get('password'))
    subjects = ['read', 'math2', 'english2','physical', 'chemisty','life' , 'japanese']
    subject_values = [float(request.form.get(subject)) if request.form.get(subject).strip() else None for subject in subjects]
    korean_geo = east_history = eco = law = None

    for subject in subjects:
        value = request.form.get(subject)
        if value.strip():
            globals()[subject] = float(value)
        else:
            globals()[subject] = None
            
    if not all(validate_score(score) for score in subject_values):
        flash('성적은 0부터 100 사이어야 합니다.')
    else:
        try:
            wb = load_workbook(excel_file_2)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I'])
        duplicate_names = [str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True)]
        if str(name) in duplicate_names:
            flash('이미 중복된 이름이 존재합니다.')
        else:
            column_headers = ['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I']
            ws.append([name, password])
            subjects = [read, math2, english2,physical, chemisty, life, korean_geo, east_history, eco, law, japanese]
            i = 3
            for subjects in subjects:
                if subjects is not None:
                    ws.cell(row=ws.max_row,column=i,value=subjects)
                i += 1
            flash('성적이 저장되었습니다.')

        wb.save(excel_file_2)

    return redirect(url_for('add_student_form_2_science'))

def add_student_2_lib():
    name = hasher(request.form.get('name'))
    password = hasher(request.form.get('password'))
    subjects = ['read', 'math2', 'english2', 'korean_geo', 'east_history', 'eco', 'law', 'japanese']
    subject_values = [float(request.form.get(subject)) if request.form.get(subject).strip() else None for subject in subjects]
    physical = chemisty = life = None
    for subject in subjects:
        value = request.form.get(subject)
        if value.strip():
            globals()[subject] = float(value)
        else:
            globals()[subject] = None
            
    if not all(validate_score(score) for score in subject_values):
        flash('성적은 0부터 100 사이어야 합니다.')
    else:
        try:
            wb = load_workbook(excel_file_2)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I'])

        duplicate_names = [str(row[0]) for row in ws.iter_rows(values_only=True)]
        if str(name) in duplicate_names:
            flash('이미 중복된 이름이 존재합니다.')
        else:
            column_headers = ['이름', '비밀번호', '독서', '수학II', '영어II','물리학I','화학I','생명I','한국지리','동아시아사','경제', '정치와 법', '일본어 I']
            ws.append([name, password])
            subjects = [read, math2, english2,physical, chemisty, life, korean_geo, east_history, eco, law, japanese]
            i = 3
            for subjects in subjects:
                if subjects is not None:
                    ws.cell(row=ws.max_row,column=i,value=subjects)
                i += 1
            flash('성적이 저장되었습니다.')

        wb.save(excel_file_2)

    return redirect(url_for('add_student_form_2_lib'))

def add_student():
    name = hasher(request.form.get('name'))
    password = hasher(request.form.get('password'))
    subjects = ['korean', 'math', 'english', 'social', 'science', 'korean_history']
    subject_values = [float(request.form.get(subject)) if request.form.get(subject).strip() else None for subject in subjects]
    for subject in subjects:
        value = request.form.get(subject)
        if value.strip():
            globals()[subject] = float(value)
        else:
            globals()[subject] = None
            
    if not all(validate_score(score) for score in subject_values):
        flash('성적은 0부터 100 사이어야 합니다.')
        return redirect(url_for('add_student_form_1'))
    else:
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['이름', '비밀번호', '국어', '수학', '영어','사회','과학','한국사'])

        duplicate_names = [str(row[0]) for row in ws.iter_rows(values_only=True)]
        if str(name) in duplicate_names:
            flash('이미 중복된 이름이 존재합니다.')
        else:
            column_headers = ["이름", "비밀번호", "국어", "수학", "영어", "사회", "과학", "한국사"]
            ws.append([name, password])
            subjects = [korean, math, english, social, science, korean_history]
            i = 3
            for subjects in subjects:
                if subjects is not None:
                    ws.cell(row=ws.max_row,column=i,value=subjects)
                i += 1
            flash('성적이 저장되었습니다.')

        wb.save(excel_file)

    return redirect(url_for('add_student_form_1'))

@app.route('/average1')
def show_cut_scores():
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        return render_template('average.html', error='데이터가 없습니다.')

    data = list(ws.iter_rows(values_only=True))
    max_score = 100  # 만점 점수

    korean_scores = [row[2] for row in data[1:]]
    math_scores = [row[3] for row in data[1:]]
    english_scores = [row[4] for row in data[1:]]
    social_scores = [row[5] for row in data[1:]]
    science_scores = [row[6] for row in data[1:]]
    korean_history_scores = [row[7] for row in data[1:]]

    # Calculate score ranges
    score_ranges = calculate_score_range([
        korean_scores, math_scores, english_scores,
        social_scores, science_scores, korean_history_scores
    ])

    return render_template('average.html',
                           korean=score_ranges[0],
                           math=score_ranges[1],
                           english=score_ranges[2],
                           social=score_ranges[3],
                           science=score_ranges[4],
                           korean_history=score_ranges[5])

@app.route('/average2')
def show_cut_scores_2():
    try:
        wb = load_workbook(excel_file_2)
        ws = wb.active
    except FileNotFoundError:
        return render_template('average_2.html', error='데이터가 없습니다.')

    data = list(ws.iter_rows(values_only=True))
    max_score = 100  # 만점 점수

    korean_scores = [row[2] for row in data[1:]]
    math_scores = [row[3] for row in data[1:]]
    english_scores = [row[4] for row in data[1:]]
    physcial_scores = [row[5] for row in data[1:]]
    chemisty_scores = [row[6] for row in data[1:]]
    life_scores = [row[7] for row in data[1:]]
    korean_geo_scores = [row[8] for row in data[1:]]
    east_history_scores = [row[9] for row in data[1:]]
    eco_scores = [row[10] for row in data[1:]]
    law_scores = [row[11] for row in data[1:]]
    japanese_scores = [row[12] for row in data[1:]]

    # Calculate score ranges
    score_ranges = calculate_score_range([
        korean_scores, math_scores, english_scores,
        physcial_scores, chemisty_scores, life_scores,korean_geo_scores,east_history_scores,
        eco_scores, law_scores,japanese_scores
    ])

    return render_template('average_2.html',
                           korean=score_ranges[0],
                           math=score_ranges[1],
                           english=score_ranges[2],
                           physical=score_ranges[3],
                           chemisty=score_ranges[4],
                           life=score_ranges[5],
                           korean_geo=score_ranges[6],
                           east_history=score_ranges[7],
                           eco=score_ranges[8],
                           law=score_ranges[9],
                           japanese=score_ranges[10])

if __name__ == '__main__':
    app.run(host="0.0.0.0", port="80", debug=False)
